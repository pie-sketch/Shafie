import pandas as pd
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import csv

# === Define Category Keywords ===
building_keywords = [
    "pangsapuri", "apartment", "apartmen", "condominium", "kondominium", "kondo", "flat", "residensi", "residence",
    "soho", "sofo", "sovo", "suite", "duplex", "loft", "villa", "townhouse", "rumah pangsa", "rumah mampu milik",
    "perumahan", "projek perumahan rakyat", "ppr", "blok", "tingkat", "unit", "lot", "floor", "wing", "cluster"
]

education_keywords = [
    "university", "kolej", "college", "sekolah", "school", "polytechnic", "politeknik", "institut", "institute",
    "international school", "maktab", "utm", "ukm", "usm", "um", "uitm", "unisel", "uoa", "ucsi", "taylor",
    "sunway", "kdu", "mahsa", "uniten", "smk", "sk"
]

mall_keywords = [
    "mall", "shopping mall", "shopping centre", "shopping center", "hypermarket", "supermarket", "plaza", "kompleks",
    "arcade", "retail", "central", "parade", "square", "galleria", "boutique mall", "megastore", "department store",
    "bazaar", "emporium", "marketplace", "souq", "souk", "avenue", "gateway", "pavilion", "ioi", "mid valley",
    "sunway pyramid", "the curve", "suria", "1 utama", "klcc", "sogo", "leisure mall", "the mines"
]

hospital_keywords = [
    "hospital", "klinik", "clinic", "medical", "pusat kesihatan", "medical centre", "health centre", "klinik desa"
]

extra_keywords = [
    "hostel", "hotel", "jabatan", "balai", "terminal", "station", "resort", "guesthouse", "dorm", "asrama"
]

# === Utility Functions ===
def is_noisy(name):
    name = name.lower().strip()
    return bool(re.fullmatch(r"jalan\s*\d+[a-z]?", name)) or ("jalan " in name and len(name.split()) <= 3)

def extract_clean_names(keys, keyword_list):
    found = set()
    for key in keys:
        key_lower = key.lower()
        if any(re.search(rf"\b{re.escape(k)}\b", key_lower) for k in keyword_list):
            if not is_noisy(key):
                found.add(key.strip())
    return sorted(found)

# === Main Function ===
def write_key_results(filepath="C:/Users/User/Desktop/Key_Check_List.xlsx"):
    df = pd.read_excel(filepath, sheet_name="Reference")
    keys = df.iloc[:, 0].dropna().astype(str).tolist()
    counter = Counter(keys)

    most_key, _ = counter.most_common(1)[0]
    duplicates = [k for k, v in counter.items() if v > 1]

    buildings = extract_clean_names(keys, building_keywords)
    educations = extract_clean_names(keys, education_keywords)
    malls = extract_clean_names(keys, mall_keywords)
    hospitals = extract_clean_names(keys, hospital_keywords)
    extras = extract_clean_names(keys, extra_keywords)

    max_len = max(len(buildings), len(educations), len(malls), len(hospitals), len(extras))

    # === Prepare Data Table ===
    headers = ["", "", "Building", "Education", "Mall", "Hospital", "Extra name 5", "Extra name 6", "Extra name 7", "Extra name 8"]
    table_data = []

    # Summary rows
    table_data.append(headers)
    table_data.append(["Most key use", most_key])
    table_data.append(["Duplicate key", ", ".join(duplicates[:5]) if duplicates else "-"])
    table_data.append(["number of duplicate", len(duplicates)])
    table_data.append(["list key duplicate", "; ".join(duplicates)])

    # Category values
    for i in range(max_len):
        row = ["", ""]
        row.append(buildings[i] if i < len(buildings) else "")
        row.append(educations[i] if i < len(educations) else "")
        row.append(malls[i] if i < len(malls) else "")
        row.append(hospitals[i] if i < len(hospitals) else "")
        row.append(extras[i] if i < len(extras) else "")
        table_data.append(row)

    # === Write to Excel ===
    wb = load_workbook(filepath)
    if "List Key" in wb.sheetnames:
        ws = wb["List Key"]
    else:
        ws = wb.create_sheet("List Key")

    # Clear existing
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
        for cell in row:
            cell.value = None

    # Fill Excel cells
    for row_idx, row_data in enumerate(table_data, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = val

    wb.save(filepath)
    print("✅ Excel file updated: 'List Key' sheet written.")

    # === Write to CSV ===
    csv_path = "C:/Users/User/Desktop/List_Key.csv"
    with open(csv_path, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in table_data:
            writer.writerow(row)

    print(f"✅ CSV file saved to: {csv_path}")

# === Run It ===
write_key_results("C:/Users/User/Desktop/Key_Check_List.xlsx")
